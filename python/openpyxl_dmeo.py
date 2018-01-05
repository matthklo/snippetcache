#!/usr/bin/env python
# -*- coding:utf-8 -*-

"""
   此篇 demo script 演示如何操作 openpyxl
   openpyxl 是一個 Python 的開源函式庫，它可以用來讀寫 Office Excel 2010 (含) 以後產出的 .xlsx 文件 (是一個開源的格式)
   
   安裝 openpyxl:
     無論在 Windows / Unix 環境下都是直接透過 pip 來安裝是最快的
   
     pip install openpyxl
   
   
   openpyxl 線上文件:
     https://openpyxl.readthedocs.io/en/default/
"""

import sys
import os

import openpyxl

if __name__ == '__main__':
  
  # 執行時需要指定一個 xlsx 檔案
  if len(sys.argv) < 2:
    print('Usage: ' + sys.argv[0] + ' {path to xlsx file}')
    os.exit(1)

  #### 讀
  
  # Workbook 就是整個 xlsx 的內容，內含一至多個 sheets，每個 sheet 裡面有很多 cells
  wb = openpyxl.load_workbook(sys.argv[1])
  
  # 查看 Workbook 裡有哪些 sheets (返回字串 list)
  sheetnames = wb.sheetnames     # sheetnames : ['工作表1', '工作表2', '工作表3']
  print('Sheet Names = ' + str(sheetnames))
  
  # 可用 sheetname 來查找 sheet 物件
  ws = wb[sheetnames[0]]
  
  # 也可用 active 取得作用中的 sheet 物件 (打開 Excel 預設會顯示的 sheet)，可能為 None
  ws = wb.active
  
  # 對 sheet 直接用 [] 操作符可提取 cell 內容
  # [] 裡面直些給 Excel 中的座標
  print('Value of cell B4 = ' + ws['B4'].value)
  
  ##### 寫
  # 待補充